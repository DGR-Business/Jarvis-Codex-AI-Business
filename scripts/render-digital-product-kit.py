import csv
import hashlib
import json
import math
import re
import shutil
import sys
import tempfile
import textwrap
import zipfile
from datetime import date, datetime
from html import escape
from pathlib import Path, PurePosixPath, PureWindowsPath

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table as WorkbookTable, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont
import pypdfium2 as pdfium
from reportlab import rl_config
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle


INK = "172033"
MUTED = "5F6B7A"
GREEN = "35C78A"
CYAN = "39B8E8"
AMBER = "F0B44D"
PANEL = "EEF2F6"
WHITE = "FFFFFF"
GRID = "D7DEE7"
FIXED_DOCUMENT_TIME = datetime(2020, 1, 1, 0, 0, 0)
FIXED_ZIP_TIME = (2020, 1, 1, 0, 0, 0)
rl_config.invariant = True


def slug(value, fallback="product", maximum=64):
    text = re.sub(r"[^a-z0-9]+", "-", str(value or "").lower()).strip("-")
    return (text or fallback)[:maximum]


def safe_text(value, maximum=500):
    return re.sub(r"\s+", " ", str(value or "")).strip()[:maximum]


WINDOWS_RESERVED_LEAF_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


def safe_output_leaf(value, label, required_suffix):
    text = str(value or "").strip()
    windows_path = PureWindowsPath(text)
    posix_path = PurePosixPath(text)
    reserved_stem = text.split(".", 1)[0].rstrip(" .").upper()
    if (
        not text
        or text in {".", ".."}
        or "/" in text
        or "\\" in text
        or windows_path.is_absolute()
        or windows_path.drive
        or posix_path.is_absolute()
        or windows_path.name != text
        or posix_path.name != text
        or text.endswith((" ", "."))
        or re.search(r'[\x00-\x1f<>:"|?*]', text)
        or reserved_stem in WINDOWS_RESERVED_LEAF_NAMES
        or Path(text).suffix.lower() != required_suffix
    ):
        raise ValueError(
            f"{label} must be a safe leaf filename ending in {required_suffix}"
        )
    return text


def contained_output_leaf(root, filename, label):
    resolved_root = root.resolve()
    candidate = (resolved_root / filename).resolve()
    if candidate.parent != resolved_root:
        raise ValueError(f"{label} escapes the renderer output directory")
    return candidate


def safe_contained_output_leaf(root, value, label, required_suffix):
    return contained_output_leaf(
        root,
        safe_output_leaf(value, label, required_suffix),
        label,
    )


def field_reference(value):
    return re.sub(r"\s+", " ", safe_text(value, 80).replace("%", " percent ")).strip().lower()


def compact_formula_evidence(formula_facts, maximum_samples=50):
    if len(formula_facts) <= maximum_samples:
        samples = list(formula_facts)
        sample_policy = "complete"
    else:
        half = maximum_samples // 2
        samples = formula_facts[:half] + formula_facts[-half:]
        sample_policy = f"first_{half}_and_last_{half}"

    coverage = []
    grouped = {}
    for fact in formula_facts:
        match = re.fullmatch(r"([A-Z]+)(\d+)", fact["cell"])
        if not match:
            coverage.append({
                "sheet": fact["sheet"],
                "range": fact["cell"],
                "count": 1,
                "firstFormula": fact["formula"],
                "lastFormula": fact["formula"],
            })
            continue
        column, row = match.group(1), int(match.group(2))
        grouped.setdefault((fact["sheet"], column), []).append((row, fact))

    for (sheet, _column), entries in grouped.items():
        entries.sort(key=lambda entry: entry[0])
        current = [entries[0]]
        runs = []
        for entry in entries[1:]:
            if entry[0] == current[-1][0] + 1:
                current.append(entry)
            else:
                runs.append(current)
                current = [entry]
        runs.append(current)
        for run in runs:
            first = run[0][1]
            last = run[-1][1]
            coverage.append({
                "sheet": sheet,
                "range": first["cell"] if len(run) == 1 else f'{first["cell"]}:{last["cell"]}',
                "count": len(run),
                "firstFormula": first["formula"],
                "lastFormula": last["formula"],
            })

    coverage.sort(key=lambda item: (item["sheet"], item["range"]))
    return {
        "samples": samples,
        "samplePolicy": sample_policy,
        "coverage": coverage,
    }


def write_json(path, value):
    path.write_text(json.dumps(value, indent=2, ensure_ascii=True), encoding="utf-8")


def canonical_zip_info(name):
    info = zipfile.ZipInfo(name, date_time=FIXED_ZIP_TIME)
    info.compress_type = zipfile.ZIP_DEFLATED
    info.create_system = 3
    info.external_attr = 0o100644 << 16
    return info


def canonicalize_zip_file(path):
    temporary = path.with_suffix(f"{path.suffix}.canonical")
    with zipfile.ZipFile(path, "r") as source:
        entries = [
            (entry.filename, source.read(entry.filename))
            for entry in source.infolist()
            if not entry.is_dir()
        ]
    with zipfile.ZipFile(temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as target:
        for name, content in sorted(entries, key=lambda item: item[0]):
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2020-01-01T00:00:00Z\g<2>",
                    content,
                )
            target.writestr(canonical_zip_info(name), content)
    temporary.replace(path)


def font(size, bold=False):
    candidates = [
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


def text_width(draw, value, selected_font):
    left, _, right, _ = draw.textbbox((0, 0), value or " ", font=selected_font)
    return right - left


def ellipsize(draw, value, selected_font, maximum_width):
    text = safe_text(value, 500)
    if text_width(draw, text, selected_font) <= maximum_width:
        return text
    suffix = "..."
    while text and text_width(draw, f"{text}{suffix}", selected_font) > maximum_width:
        text = text[:-1].rstrip()
    return f"{text}{suffix}" if text else suffix


def wrapped_lines(draw, value, selected_font, maximum_width, maximum_lines):
    words = safe_text(value, 1000).split()
    lines = []
    current = ""
    while words:
        word = words.pop(0)
        candidate = f"{current} {word}".strip()
        if text_width(draw, candidate, selected_font) <= maximum_width:
            current = candidate
            continue
        if current:
            lines.append(current)
            current = ""
            words.insert(0, word)
        else:
            chunk = ""
            for character in word:
                candidate = f"{chunk}{character}"
                if chunk and text_width(draw, candidate, selected_font) > maximum_width:
                    lines.append(chunk)
                    chunk = character
                else:
                    chunk = candidate
            current = chunk
        if len(lines) >= maximum_lines:
            break
    if current and len(lines) < maximum_lines:
        lines.append(current)
    remaining = bool(words) or (current and len(lines) >= maximum_lines and lines[-1] != current)
    if remaining and lines:
        lines[-1] = ellipsize(draw, f"{lines[-1]} ...", selected_font, maximum_width)
    return lines[:maximum_lines] or [""]


def fitted_wrapped_font(draw, value, maximum_width, maximum_lines, starting_size, minimum_size, bold=False):
    for size in range(starting_size, minimum_size - 1, -1):
        selected = font(size, bold)
        lines = wrapped_lines(draw, value, selected, maximum_width, 1000)
        if len(lines) <= maximum_lines:
            return selected, lines
    selected = font(minimum_size, bold)
    return selected, wrapped_lines(draw, value, selected, maximum_width, maximum_lines)


def draw_wrapped(draw, position, value, fill, selected_font, maximum_width, maximum_lines, spacing=4):
    lines = wrapped_lines(draw, value, selected_font, maximum_width, maximum_lines)
    x, y = position
    draw.multiline_text((x, y), "\n".join(lines), fill=fill, font=selected_font, spacing=spacing)
    line_height = draw.textbbox((0, 0), "Ag", font=selected_font)[3]
    return y + len(lines) * line_height + max(0, len(lines) - 1) * spacing


def parse_cell(value, kind):
    text = safe_text(value, 240)
    literal = f"'{text}" if text.startswith(("=", "+", "-", "@")) else text
    if kind == "date":
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            return literal
    if kind in {"number", "currency", "percent"}:
        cleaned = re.sub(r"[^0-9.\-]", "", text)
        try:
            number = float(cleaned)
            return number / 100 if kind == "percent" and "%" in text else number
        except ValueError:
            return literal
    if kind == "boolean":
        return "Yes" if text.lower() in {"yes", "true", "1", "done", "complete"} else "No"
    return literal


def json_cell(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def claim_safe_copy(value):
    text = safe_text(value, 2000)
    replacements = [
        (
            "collect better inputs",
            "collect project inputs in a structured format",
        ),
        (
            "Close projects cleanly with organized files, usage notes, final approvals, and a next-step invitation.",
            "Track file-index checks, usage notes, final approvals, and next steps in one place.",
        ),
    ]
    for source, replacement in replacements:
        text = re.sub(re.escape(source), replacement, text, flags=re.IGNORECASE)
    return text


RUNTIME_TOPOLOGY_REFERENCE = re.compile(
    r"\b(?:workbooks?|worksheets?|spreadsheets?|sheets?|tabs?|files?|folders?|archives?|zips?|csvs?|pdfs?)\b"
    r"|\.(?:xlsx|csv|pdf|zip)\b",
    flags=re.IGNORECASE,
)
UNVERIFIED_NAMED_STRUCTURE = re.compile(
    r"\b(?:to|in|from|within|into|open|update|use|maintain)\s+(?:the\s+)?"
    r"(?:[A-Za-z][A-Za-z0-9&/-]*\s+){0,3}"
    r"(?:register|matrix|board|records?|log|index|databases?|portal)\b",
    flags=re.IGNORECASE,
)


def canonical_package_setup_steps(catalogue_items):
    if len(catalogue_items) == 1:
        workbook_name = f"01-{slug(catalogue_items[0].get('title'))}.xlsx"
        open_instruction = (
            f"Open the setup guide, then open the single included workbook: {workbook_name}."
        )
    else:
        open_instruction = (
            "Open the setup guide, then choose the workbook that matches the job you are doing."
        )
    return [
        "Download and unzip the customer bundle into a working folder.",
        open_instruction,
        "Read the workbook's Read Me sheet before replacing its example records.",
        "Use reviewed information and keep the original bundle as a clean backup.",
        "Review the Dashboard before using or sharing the updated records.",
    ]


def claim_safe_item_purpose(item):
    purpose = claim_safe_copy(item.get("purpose"))
    field_names = [safe_text(column.get("name"), 100) for column in item.get("columns", [])]
    field_text = " ".join(field_names)
    adjustments = []
    claims_content_format = re.search(r"\b(?:content\s+)?formats?\b", purpose, flags=re.IGNORECASE)
    has_content_format_field = re.search(
        r"\b(?:format|content type|asset type|post type|media type)\b",
        field_text,
        flags=re.IGNORECASE,
    )
    if claims_content_format and not has_content_format_field and field_names:
        visible_fields = field_names[:5]
        if len(visible_fields) == 1:
            field_summary = visible_fields[0]
        elif len(visible_fields) == 2:
            field_summary = " and ".join(visible_fields)
        else:
            field_summary = f"{', '.join(visible_fields[:-1])}, and {visible_fields[-1]}"
        purpose = f"Record and review {field_summary} in one editable tracker."
        adjustments.append({
            "code": "purpose_rewritten_from_actual_fields",
            "reason": "The source purpose referred to content formats, but no format field exists.",
        })
    return purpose, adjustments


def truth_aligned_item_structure(source_item):
    item = {
        **source_item,
        "columns": [dict(column) for column in source_item.get("columns", [])],
        "instructions": list(source_item.get("instructions", [])),
        "calculations": [dict(calculation) for calculation in source_item.get("calculations", [])],
    }
    adjustments = []
    renamed_fields = {}
    existing_names = {
        field_reference(column.get("name"))
        for column in item["columns"]
    }

    for column in item["columns"]:
        original_name = safe_text(column.get("name"), 100)
        if not re.fullmatch(r"client[ -]facing status", original_name, flags=re.IGNORECASE):
            continue
        replacement = "Status to Share"
        if field_reference(replacement) in existing_names and field_reference(original_name) != field_reference(replacement):
            replacement = "Shareable Status"
        renamed_fields[original_name] = replacement
        existing_names.discard(field_reference(original_name))
        existing_names.add(field_reference(replacement))
        column["name"] = replacement
        column["guidance"] = "Record a concise status that can be copied into a client update."
        adjustments.append({
            "code": "client_display_claim_narrowed",
            "reason": (
                "Pantheon changed an ambiguous client-display field into a literal status-to-share field; "
                "the workbook does not claim to provide a hosted client view."
            ),
        })

    aligned_instructions = []
    fields = {field_reference(column.get("name")) for column in item["columns"]}
    instruction_changed = False
    for source in item["instructions"]:
        instruction = safe_text(source, 500)
        aligned = instruction
        for original_name, replacement in renamed_fields.items():
            aligned = re.sub(re.escape(original_name), replacement, aligned, flags=re.IGNORECASE)
        aligned = re.sub(
            r"\bclient-facing field\b",
            "Status to Share field",
            aligned,
            flags=re.IGNORECASE,
        )
        if {"requested assets", "next milestone"}.issubset(fields):
            aligned = re.sub(
                r"\bRequested Assets and Milestones\b",
                "Requested Assets and Next Milestone",
                aligned,
                flags=re.IGNORECASE,
            )
            if re.search(
                r"\bkeep\b.*\bvisible\b.*\bclient\b",
                aligned,
                flags=re.IGNORECASE,
            ):
                aligned = (
                    "Use Requested Assets and Next Milestone to keep delivery work organized for "
                    "the project owner when preparing client updates."
                )
        if re.search(r"\bStatus to Share\b", aligned, flags=re.IGNORECASE):
            aligned = re.sub(
                r"Use Status to Share to display a concise project state without exposing private notes\.?",
                (
                    "Use Status to Share to record a concise project state that can be copied into a "
                    "client update while keeping private notes separate."
                ),
                aligned,
                flags=re.IGNORECASE,
            )
            aligned = re.sub(r"\bclient-facing view\b", "client update", aligned, flags=re.IGNORECASE)
            aligned = re.sub(r"\bclient display\b", "client update", aligned, flags=re.IGNORECASE)
        if aligned != instruction:
            instruction_changed = True
        aligned_instructions.append(aligned)
    item["instructions"] = aligned_instructions

    if renamed_fields:
        for calculation in item["calculations"]:
            target = safe_text(calculation.get("target"), 100)
            calculation["target"] = renamed_fields.get(target, target)
            calculation["inputs"] = [
                renamed_fields.get(safe_text(value, 100), safe_text(value, 100))
                for value in calculation.get("inputs", [])
            ]
    if instruction_changed:
        adjustments.append({
            "code": "instruction_field_names_aligned",
            "reason": "Pantheon aligned customer instructions to the exact fields present in the workbook.",
        })
    return item, adjustments


def model_customer_steps(item, maximum=3):
    kept = []
    removed = 0
    for source in item.get("instructions", []):
        instruction = safe_text(source, 500)
        if not instruction:
            continue
        if (
            RUNTIME_TOPOLOGY_REFERENCE.search(instruction)
            or UNVERIFIED_NAMED_STRUCTURE.search(instruction)
        ):
            removed += 1
            continue
        if instruction.casefold() in {value.casefold() for value in kept}:
            continue
        kept.append(instruction)
        if len(kept) >= maximum:
            break
    adjustments = []
    if removed:
        adjustments.append({
            "code": "model_topology_replaced",
            "reason": (
                f"Pantheon replaced {removed} model-authored file or sheet instruction(s) "
                "with instructions derived from the files it actually created."
            ),
        })
    return kept, adjustments


def status_options(column, sample_rows, column_index):
    def clean_option(value):
        # Keep dropdown values as data. Guidance after a semicolon belongs in Read Me.
        return safe_text(value, 80).split(";", 1)[0].strip()

    declared = column.get("options")
    options = [
        clean_option(value)
        for value in declared
        if clean_option(value)
    ] if isinstance(declared, list) else []
    if not options:
        guidance = safe_text(column.get("guidance"), 500)
        match = re.search(r"\bUse\s+(.+?)(?:[.!?]|$)", guidance, flags=re.IGNORECASE)
        if match:
            choices = re.sub(r"\s*,?\s+(?:or|and)\s+", ",", match.group(1), flags=re.IGNORECASE)
            options = [
                clean_option(value.strip(" '\""))
                for value in choices.split(",")
                if clean_option(value.strip(" '\""))
            ]
    sample_options = [
        clean_option(source_row[column_index])
        for source_row in sample_rows
        if len(source_row) > column_index and clean_option(source_row[column_index])
    ]
    if not options:
        options = ["Not started", "In progress", "Waiting", "Complete", "Cancelled"]
    return list(dict.fromkeys([*options, *sample_options]))


def positive_status(options):
    priorities = [
        "Complete",
        "Completed",
        "Closed",
        "Approved",
        "Ready",
        "Confirmed",
        "Accepted",
        "Sent",
        "Applied",
        "On track",
        "Archived",
    ]
    by_lower = {value.lower(): value for value in options}
    return next((by_lower[value.lower()] for value in priorities if value.lower() in by_lower), None)


def clarified_column_guidance(column):
    guidance = safe_text(column.get("guidance"), 500)
    if column.get("type") != "status":
        return guidance
    options = [safe_text(value, 80) for value in column.get("options", [])]
    by_lower = {value.lower(): value for value in options}
    if "ready" in by_lower and "approved" in by_lower:
        base = guidance.rstrip(" .")
        return (
            f"{base}. Ready means prepared for final review; Approved means accepted for use. "
            "The Dashboard counts only Approved records as complete."
        )
    return guidance


def sample_row_height(tracker, row_index, column_count):
    line_count = 1
    for column_index in range(1, column_count + 1):
        cell = tracker.cell(row=row_index, column=column_index)
        value = "" if cell.value is None else str(cell.value)
        width = float(tracker.column_dimensions[cell.column_letter].width or 14)
        characters_per_line = max(8, int(width * 0.75))
        cell_lines = sum(
            max(1, len(textwrap.wrap(
                part,
                width=characters_per_line,
                break_long_words=True,
                break_on_hyphens=True,
            )))
            for part in value.splitlines() or [""]
        )
        line_count = max(line_count, cell_lines)
    return min(158, max(34, (min(line_count, 10) * 15) + 8))


def dashboard_status_field(status_fields):
    preferred = [
        field
        for field in status_fields
        if re.search(
            r"(?:^|\s)(?:status|workflow status|message status|contact status|completion status)$",
            safe_text(field.get("field"), 80),
            flags=re.IGNORECASE,
        )
    ]
    return preferred[0] if preferred else status_fields[0]


def customer_instructions(item, workbook_name, csv_name):
    controlled_fields = [
        column["name"]
        for column in item["columns"]
        if column["type"] in {"status", "boolean"}
    ]
    controlled_note = (
        f"Use the supplied dropdown choices for {', '.join(controlled_fields)} so the records stay consistent."
        if controlled_fields
        else "Keep each field consistent so filters and counts remain useful."
    )
    calculations = item.get("calculations", [])
    calculated_note = (
        "Calculated fields update from the named input columns; do not overwrite their formulas."
        if calculations
        else None
    )
    task_steps, adjustments = model_customer_steps(item)
    instructions = [
        f"Open {workbook_name} and read the Read Me sheet before replacing any example data.",
        *task_steps,
        f"Use the Tracker sheet to {safe_text(item['purpose'], 500).lower().rstrip('.')}.",
        f"Review the {len(item['sampleRows'])} example records, then replace them with your own information or import {csv_name} into another approved tool.",
        controlled_note,
        calculated_note,
        "Check the Dashboard after updates; its record and status counts recalculate from the Tracker sheet when the workbook opens.",
    ] if calculated_note else [
        f"Open {workbook_name} and read the Read Me sheet before replacing any example data.",
        *task_steps,
        f"Use the Tracker sheet to {safe_text(item['purpose'], 500).lower().rstrip('.')}.",
        f"Review the {len(item['sampleRows'])} example records, then replace them with your own information or import {csv_name} into another approved tool.",
        controlled_note,
        "Check the Dashboard after updates; its record and status counts recalculate from the Tracker sheet when the workbook opens.",
    ]
    return list(dict.fromkeys(instruction for instruction in instructions if instruction)), adjustments


def calculation_definitions(item, headers):
    positions = {
        field_reference(name): {"index": index, "name": name}
        for index, name in enumerate(headers, start=1)
    }
    if len(positions) != len(headers):
        raise ValueError(f"Ambiguous calculated-field names in {item['id']}")
    calculations = item.get("calculations", [])
    if not isinstance(calculations, list):
        raise ValueError(f"Calculations for {item['id']} must be a list")
    definitions = []
    targets = set()
    for calculation in calculations:
        target = field_reference(calculation.get("target"))
        operation = safe_text(calculation.get("operation"), 40)
        inputs = [field_reference(value) for value in calculation.get("inputs", [])]
        if (
            target not in positions
            or target in targets
            or operation not in {"multiply", "sum", "subtract", "percent_of"}
            or len(inputs) < 2
            or len(inputs) > 6
            or any(value not in positions or value == target for value in inputs)
            or (operation in {"multiply", "subtract", "percent_of"} and len(inputs) != 2)
        ):
            raise ValueError(
                f"Invalid calculated field in {item['id']}: "
                f"{calculation.get('target') or 'unnamed'}; {operation or 'no operation'}; "
                f"{', '.join(inputs) or 'no inputs'}"
            )
        targets.add(target)
        definitions.append({
            "target": positions[target]["name"],
            "targetReference": target,
            "operation": operation,
            "inputs": [positions[value]["name"] for value in inputs],
            "inputReferences": inputs,
            "targetColumn": positions[target]["index"],
            "inputColumns": [positions[value]["index"] for value in inputs],
        })
    dependencies = {
        definition["targetReference"]: definition["inputReferences"]
        for definition in definitions
    }

    def reaches(start, target, visiting=None):
        if start == target:
            return True
        visiting = set() if visiting is None else visiting
        if start in visiting:
            return False
        visiting.add(start)
        return any(
            dependency in dependencies and reaches(dependency, target, visiting)
            for dependency in dependencies.get(start, [])
        )

    for target, inputs in dependencies.items():
        if any(value in dependencies and reaches(value, target) for value in inputs):
            raise ValueError(f"Circular calculated field in {item['id']}: {positions[target]['name']}")
    return definitions


def calculation_formula(definition, tracker, row_index):
    references = [
        f"{tracker.cell(row=row_index, column=index).column_letter}{row_index}"
        for index in definition["inputColumns"]
    ]
    if definition["operation"] == "multiply":
        expression = f"{references[0]}*{references[1]}"
    elif definition["operation"] == "sum":
        expression = f"SUM({','.join(references)})"
    elif definition["operation"] == "subtract":
        expression = f"{references[0]}-{references[1]}"
    elif definition["operation"] == "percent_of":
        return f'=IF(OR({references[0]}="",{references[1]}="",{references[1]}=0),"",{references[0]}/{references[1]})'
    else:
        raise ValueError(f"Unsupported calculation operation: {definition['operation']}")
    return f'=IF(COUNTA({",".join(references)})=0,"",{expression})'


def numeric_value(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = safe_text(value, 80).replace("A$", "").replace("$", "").replace(",", "")
    if not text:
        return None
    percent = text.endswith("%")
    if percent:
        text = text[:-1]
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100 if percent else number


def display_calculation(value, kind):
    if value is None:
        return ""
    if kind == "currency":
        return f"A${value:,.2f}"
    if kind == "percent":
        return f"{value:.0%}"
    return f"{value:,.2f}".rstrip("0").rstrip(".")


def sample_calculation_checks(item, definitions):
    headers = [safe_text(column["name"], 80) for column in item["columns"]]
    kinds = [column["type"] for column in item["columns"]]
    checks = []
    for row_index, source_row in enumerate(item["sampleRows"], start=2):
        values = {
            field_reference(header): parse_cell(source_row[index], kinds[index])
            for index, header in enumerate(headers)
        }
        results = []
        for definition in definitions:
            inputs = [
                numeric_value(values.get(field_reference(name)))
                for name in definition["inputs"]
            ]
            result = None
            if all(value is not None for value in inputs):
                if definition["operation"] == "multiply":
                    result = inputs[0] * inputs[1]
                elif definition["operation"] == "sum":
                    result = sum(inputs)
                elif definition["operation"] == "subtract":
                    result = inputs[0] - inputs[1]
                elif definition["operation"] == "percent_of" and inputs[1] != 0:
                    result = inputs[0] / inputs[1]
            target_index = definition["targetColumn"] - 1
            values[field_reference(definition["target"])] = result
            results.append({
                "target": definition["target"],
                "operation": definition["operation"],
                "inputs": definition["inputs"],
                "value": result,
                "display": display_calculation(result, kinds[target_index]),
            })
        checks.append({
            "trackerRow": row_index,
            "record": safe_text(source_row[0], 120) if source_row else f"Row {row_index}",
            "results": results,
        })
    return checks


def workbook_for_item(item, package_title, customer_promise, setup_steps, output_path):
    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read Me"
    tracker = workbook.create_sheet("Tracker")
    dashboard = workbook.create_sheet("Dashboard", 0)

    readme.sheet_view.showGridLines = False
    readme.column_dimensions["A"].width = 4
    readme.column_dimensions["B"].width = 28
    readme.column_dimensions["C"].width = 82
    readme.merge_cells("B2:C2")
    readme["B2"] = safe_text(item["title"], 160)
    readme["B2"].font = Font(name="Aptos Display", size=22, bold=True, color=INK)
    readme["B2"].alignment = Alignment(vertical="center", wrap_text=True)
    readme.row_dimensions[2].height = 36
    readme["B4"] = "Purpose"
    readme["C4"] = safe_text(item["purpose"], 1000)
    readme["B6"] = "How to use it"
    for offset, instruction in enumerate(item["instructions"], start=6):
        readme[f"C{offset}"] = f"{offset - 5}. {safe_text(instruction, 500)}"
    guide_start = 7 + len(item["instructions"])
    readme[f"B{guide_start}"] = "Field guide"
    for index, column in enumerate(item["columns"], start=guide_start):
        readme[f"C{index}"] = f"{safe_text(column['name'], 80)}: {safe_text(column['guidance'], 300)}"
    for cell in [f"B{row}" for row in [4, 6, guide_start]]:
        readme[cell].font = Font(bold=True, color=GREEN)
    for row in readme.iter_rows(min_row=1, max_row=readme.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    headers = [safe_text(column["name"], 80) for column in item["columns"]]
    kinds = [column["type"] for column in item["columns"]]
    calculations = calculation_definitions(item, headers)
    sample_checks = sample_calculation_checks(item, calculations)
    for index, header in enumerate(headers, start=1):
        cell = tracker.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        tracker.column_dimensions[cell.column_letter].width = min(30, max(14, len(header) + 4))
    tracker.row_dimensions[1].height = 36
    for row_index, source_row in enumerate(item["sampleRows"], start=2):
        for column_index, value in enumerate(source_row, start=1):
            kind = kinds[column_index - 1]
            cell = tracker.cell(row=row_index, column=column_index, value=parse_cell(value, kind))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if kind == "date":
                cell.number_format = "dd mmm yyyy"
            elif kind == "currency":
                cell.number_format = 'A$#,##0.00'
            elif kind == "percent":
                cell.number_format = "0%"
        tracker.row_dimensions[row_index].height = sample_row_height(
            tracker,
            row_index,
            len(headers),
        )
    tracker.freeze_panes = "A2"
    tracker.auto_filter.ref = f"A1:{tracker.cell(row=max(2, tracker.max_row), column=len(headers)).coordinate}"
    table_end_row = 500 if calculations else max(2, tracker.max_row)
    table_ref = f"A1:{tracker.cell(row=table_end_row, column=len(headers)).coordinate}"
    table = WorkbookTable(displayName=f"Tracker_{slug(item['id'], 'item', 24).replace('-', '_')}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tracker.add_table(table)
    thin = Side(style="thin", color=GRID)
    for row in tracker.iter_rows(min_row=1, max_row=max(2, tracker.max_row), max_col=len(headers)):
        for cell in row:
            cell.border = Border(bottom=thin)

    status_fields = []
    for column_index, kind in enumerate(kinds, start=1):
        letter = tracker.cell(row=1, column=column_index).column_letter
        if kind == "status":
            options = status_options(item["columns"][column_index - 1], item["sampleRows"], column_index - 1)
            if any("," in option for option in options):
                raise ValueError(f"Status choices for {headers[column_index - 1]} cannot contain commas")
            if len(",".join(options)) > 240:
                raise ValueError(f"Status choices for {headers[column_index - 1]} exceed the workbook dropdown limit")
            status_fields.append({
                "field": headers[column_index - 1],
                "column": letter,
                "options": options,
                "positiveStatus": positive_status(options),
            })
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(options)}"',
                allow_blank=True,
            )
            tracker.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}500")
            if status_fields[-1]["positiveStatus"]:
                tracker.conditional_formatting.add(
                    f"{letter}2:{letter}500",
                    FormulaRule(
                        formula=[f'{letter}2="{status_fields[-1]["positiveStatus"]}"'],
                        fill=PatternFill("solid", fgColor="DDF5EA"),
                    ),
                )
        elif kind == "boolean":
            validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            tracker.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}500")

    calculated_fields = []
    for definition in calculations:
        target_kind = kinds[definition["targetColumn"] - 1]
        target_letter = tracker.cell(row=1, column=definition["targetColumn"]).column_letter
        for row_index in range(2, 501):
            cell = tracker.cell(
                row=row_index,
                column=definition["targetColumn"],
                value=calculation_formula(definition, tracker, row_index),
            )
            if target_kind == "currency":
                cell.number_format = 'A$#,##0.00'
            elif target_kind == "percent":
                cell.number_format = "0%"
            elif target_kind == "number":
                cell.number_format = "0.00"
        calculated_fields.append({
            "target": definition["target"],
            "operation": definition["operation"],
            "inputs": definition["inputs"],
            "formula": tracker.cell(row=2, column=definition["targetColumn"]).value,
            "range": f"{target_letter}2:{target_letter}500",
        })

    dashboard.sheet_view.showGridLines = False
    dashboard.column_dimensions["A"].width = 4
    dashboard.column_dimensions["B"].width = 28
    dashboard.column_dimensions["C"].width = 26
    dashboard.merge_cells("B2:C2")
    dashboard["B2"] = safe_text(item["title"], 160)
    dashboard["B2"].font = Font(name="Aptos Display", size=19, bold=True, color=INK)
    dashboard["B2"].alignment = Alignment(vertical="center", wrap_text=True)
    dashboard.row_dimensions[2].height = 48
    dashboard["B4"] = "Records entered"
    dashboard["C4"] = "=COUNTA('Tracker'!A2:A500)"
    dashboard["B5"] = "Fields per record"
    dashboard["C5"] = len(headers)
    dashboard["B7"] = "Use this dashboard"
    dashboard["C7"] = "Add or update records in the Tracker sheet. The count above updates automatically."
    dashboard_metric = None
    if status_fields:
        status_field = dashboard_status_field(status_fields)
        if status_field["positiveStatus"]:
            dashboard["B6"] = f'{status_field["positiveStatus"]} records'
            dashboard["C6"] = (
                f'=COUNTIF(\'Tracker\'!{status_field["column"]}2:{status_field["column"]}500,'
                f'"{status_field["positiveStatus"]}")'
            )
            dashboard_metric = {
                "label": dashboard["B6"].value,
                "formula": dashboard["C6"].value,
                "statusField": status_field["field"],
                "countedValue": status_field["positiveStatus"],
                "countedValueInValidation": status_field["positiveStatus"] in status_field["options"],
            }
        else:
            dashboard["B6"] = f'{status_field["field"]} recorded'
            dashboard["C6"] = f'=COUNTA(\'Tracker\'!{status_field["column"]}2:{status_field["column"]}500)'
            dashboard_metric = {
                "label": dashboard["B6"].value,
                "formula": dashboard["C6"].value,
                "statusField": status_field["field"],
                "countedValue": None,
                "countedValueInValidation": True,
            }
    for row in range(4, 8):
        dashboard[f"B{row}"].font = Font(bold=True, color=GREEN if row in {4, 5, 6} else INK)
        dashboard[f"C{row}"].alignment = Alignment(wrap_text=True)
    dashboard["B10"] = safe_text(item["purpose"], 600)
    dashboard["B10"].alignment = Alignment(wrap_text=True)
    dashboard.merge_cells("B10:C12")
    dashboard["B10"].fill = PatternFill("solid", fgColor=PANEL)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.properties.creator = "Pantheon"
    workbook.properties.lastModifiedBy = "Pantheon"
    workbook.properties.created = FIXED_DOCUMENT_TIME
    workbook.properties.modified = FIXED_DOCUMENT_TIME
    workbook.save(output_path)
    canonicalize_zip_file(output_path)
    reopened = load_workbook(output_path, data_only=False, read_only=False)
    required_sheets = {"Dashboard", "Tracker", "Read Me"}
    if set(reopened.sheetnames) != required_sheets:
        raise ValueError(f"Workbook {output_path.name} did not reopen with the required sheets")
    all_formula_facts = [
        {
            "sheet": sheet.title,
            "cell": cell.coordinate,
            "formula": cell.value,
        }
        for sheet in reopened.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    formula_evidence = compact_formula_evidence(all_formula_facts)
    tracker_sheet = reopened["Tracker"]
    readme_sheet = reopened["Read Me"]
    reopened_headers = [
        tracker_sheet.cell(row=1, column=index).value
        for index in range(1, len(headers) + 1)
    ]
    reopened_rows = [
        [
            json_cell(tracker_sheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, len(headers) + 1)
        ]
        for row_index in range(2, 2 + len(item["sampleRows"]))
    ]
    instruction_facts = [
        {
            "cell": f"Read Me!C{row_index}",
            "text": readme_sheet[f"C{row_index}"].value,
        }
        for row_index in range(6, 6 + len(item["instructions"]))
    ]
    field_facts = [
        {
            "name": column["name"],
            "type": column["type"],
            "guidance": column["guidance"],
            "trackerHeader": reopened_headers[index],
            "readMeCell": f"Read Me!C{guide_start + index}",
            "readMeText": readme_sheet[f"C{guide_start + index}"].value,
        }
        for index, column in enumerate(item["columns"])
    ]
    validation_facts = [
        {
            "sheet": "Tracker",
            "type": validation.type,
            "formula": validation.formula1,
            "range": str(validation.sqref),
            "allowBlank": validation.allow_blank is True,
        }
        for validation in tracker_sheet.data_validations.dataValidation
    ]
    reopened.close()
    return {
        "sheets": sorted(required_sheets),
        "columns": len(headers),
        "sampleRows": len(item["sampleRows"]),
        "formulaCells": len(all_formula_facts),
        "reopened": True,
        "instructions": instruction_facts,
        "fields": field_facts,
        "sampleData": {
            "headers": reopened_headers,
            "rows": reopened_rows,
        },
        "formulas": formula_evidence["samples"],
        "formulaSamplePolicy": formula_evidence["samplePolicy"],
        "formulaCoverage": formula_evidence["coverage"],
        "calculatedFields": calculated_fields,
        "sampleCalculationChecks": sample_checks,
        "dashboardExpectedResults": {
            "recordsEntered": sum(
                1 for row in item["sampleRows"]
                if row and safe_text(row[0], 120)
            ),
            "metricLabel": dashboard_metric["label"] if dashboard_metric else None,
            "metricValue": (
                sum(
                    1 for row in item["sampleRows"]
                    if dashboard_metric
                    and safe_text(
                        row[next(
                            index for index, header in enumerate(headers)
                            if header == dashboard_metric["statusField"]
                        )],
                        80,
                    ) == dashboard_metric["countedValue"]
                )
                if dashboard_metric and dashboard_metric["countedValue"]
                else None
            ),
        },
        "dataValidations": validation_facts,
        "statusFields": status_fields,
        "dashboardMetric": dashboard_metric,
        "sheetSummary": {
            "Dashboard": "Live record and completion counts derived from the Tracker sheet.",
            "Tracker": "Editable customer records with filters, formatting, sample rows, and controlled fields.",
            "Read Me": "Product purpose, customer instructions, and field-by-field guidance.",
        },
    }


def write_csv(item, output_path):
    headers = [safe_text(column["name"], 80) for column in item["columns"]]
    kinds = [column["type"] for column in item["columns"]]
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows([
            [
                json_cell(parse_cell(value, kinds[index]))
                for index, value in enumerate(row)
            ]
            for row in item["sampleRows"]
        ])


def guide_pdf(blueprint, customer_files, output_path):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PantheonTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=21,
        leading=25,
        textColor=colors.HexColor(f"#{INK}"),
        alignment=TA_LEFT,
        spaceAfter=10,
    )
    heading = ParagraphStyle(
        "PantheonHeading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
        textColor=colors.HexColor(f"#{INK}"),
        spaceBefore=10,
        spaceAfter=6,
    )
    body = ParagraphStyle(
        "PantheonBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.2,
        leading=13,
        textColor=colors.HexColor(f"#{MUTED}"),
        spaceAfter=6,
    )
    compact_body = ParagraphStyle(
        "PantheonCompactBody",
        parent=body,
        fontSize=8.8,
        leading=11.5,
        spaceAfter=3.5,
    )
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=safe_text(blueprint["packageTitle"], 180),
    )
    story = [
        Paragraph(escape(safe_text(blueprint["packageTitle"], 180)), title_style),
        Paragraph(escape(safe_text(blueprint["customerPromise"], 900)), body),
        Spacer(1, 5 * mm),
        Paragraph("Quick start", heading),
    ]
    for index, step in enumerate(blueprint["setupSteps"], start=1):
        story.append(Paragraph(f"<b>{index}.</b> {escape(safe_text(step, 700))}", body))
    story.extend([Spacer(1, 4 * mm), Paragraph("What is included", heading)])
    table_data = [["Product", "Files"]]
    for item in blueprint["catalogueItems"]:
        paths = customer_files[item["id"]]
        table_data.append([
            Paragraph(escape(safe_text(item["title"], 120)), body),
            Paragraph("<br/>".join(escape(Path(value).name) for value in paths), body),
        ])
    table = PdfTable(table_data, colWidths=[70 * mm, 80 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{INK}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(f"#{GRID}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{PANEL}")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    for item in blueprint["catalogueItems"]:
        story.extend([
            PageBreak(),
            Paragraph(escape(safe_text(item["title"], 180)), title_style),
            Paragraph(escape(safe_text(item["purpose"], 1000)), body),
            Paragraph("Recommended workflow", heading),
        ])
        for index, instruction in enumerate(item["instructions"], start=1):
            story.append(Paragraph(f"<b>{index}.</b> {escape(safe_text(instruction, 700))}", body))
        story.append(Paragraph("Fields", heading))
        for column in item["columns"]:
            story.append(Paragraph(
                f"<b>{escape(safe_text(column['name'], 100))}</b> - {escape(safe_text(column['guidance'], 500))}",
                compact_body,
            ))
    if blueprint["disclaimers"]:
        notes = [Spacer(1, 3 * mm), Paragraph("Important notes", heading)]
        for disclaimer in blueprint["disclaimers"]:
            notes.append(Paragraph(escape(safe_text(disclaimer, 800)), compact_body))
        story.append(KeepTogether(notes))
    doc.build(story)


def preview_image(blueprint, output_path, mode):
    image = Image.new("RGB", (1400, 900), f"#{INK}")
    draw = ImageDraw.Draw(image)
    draw.rectangle((50, 50, 1350, 850), fill="#F7F9FC")
    title_font, title_lines = fitted_wrapped_font(
        draw,
        blueprint["packageTitle"],
        1210,
        2,
        42,
        32,
        True,
    )
    draw.multiline_text((95, 82), "\n".join(title_lines), fill=f"#{INK}", font=title_font, spacing=5)
    title_height = draw.textbbox((0, 0), "Ag", font=title_font)[3]
    title_bottom = 82 + len(title_lines) * title_height + max(0, len(title_lines) - 1) * 5
    promise_bottom = draw_wrapped(
        draw,
        (95, title_bottom + 12),
        blueprint["customerPromise"],
        f"#{MUTED}",
        font(21),
        1210,
        2,
        5,
    )
    if mode == "dashboard":
        section_y = max(220, promise_bottom + 22)
        item_count = len(blueprint["catalogueItems"])
        tool_label = "tool" if item_count == 1 else "tools"
        draw.text(
            (95, section_y),
            f"{item_count} focused {tool_label} built from the real customer files.",
            fill=f"#{GREEN}",
            font=font(24, True),
        )
        if item_count == 1:
            item = blueprint["catalogueItems"][0]
            card_top = min(300, max(274, section_y + 48))
            draw.rounded_rectangle((95, card_top, 1305, card_top + 190), radius=8, fill="#FFFFFF", outline=f"#{GRID}", width=2)
            draw.rectangle((95, card_top, 107, card_top + 190), fill=f"#{GREEN}")
            item_font, item_lines = fitted_wrapped_font(draw, item["title"], 1130, 2, 26, 20, True)
            draw.multiline_text((130, card_top + 24), "\n".join(item_lines), fill=f"#{INK}", font=item_font, spacing=3)
            title_height = draw.textbbox((0, 0), "Ag", font=item_font)[3]
            purpose_y = card_top + 24 + len(item_lines) * title_height + max(0, len(item_lines) - 1) * 3 + 12
            draw_wrapped(draw, (130, purpose_y), item["purpose"], f"#{MUTED}", font(18), 1120, 3, 4)

            panel_top = card_top + 220
            panels = [
                ("Workflow control", "Work Status | Due Date | Delivery Proof", GREEN),
                ("Client decisions", "Approval Status | Scope Change | Notes", CYAN),
                ("Profit view", "Quoted Fee | Total Cost | Estimated Contribution", AMBER),
            ]
            for index, (label, detail, colour) in enumerate(panels):
                left = 95 + index * 410
                draw.rounded_rectangle((left, panel_top, left + 380, panel_top + 150), radius=8, fill="#FFFFFF", outline=f"#{GRID}", width=2)
                draw.rectangle((left, panel_top, left + 8, panel_top + 150), fill=f"#{colour}")
                draw.text((left + 28, panel_top + 24), label, fill=f"#{INK}", font=font(19, True))
                draw_wrapped(draw, (left + 28, panel_top + 62), detail, f"#{MUTED}", font(15), 325, 3, 4)
            draw.text(
                (95, min(812, panel_top + 180)),
                f"Editable Excel workbook | setup guide | sample CSV | {len(item['sampleRows'])} example records",
                fill=f"#{MUTED}",
                font=font(16),
            )
            image.save(output_path, format="PNG", optimize=True)
            with Image.open(output_path) as reopened:
                reopened.verify()
            return
        y = min(310, max(276, section_y + 52))
        colours = [GREEN, CYAN, AMBER, "8C86F7", "EC6F91", "61C0BF"]
        for index, item in enumerate(blueprint["catalogueItems"]):
            x = 95 + (index % 2) * 620
            if index and index % 2 == 0:
                y += 160
            draw.rounded_rectangle((x, y, x + 570, y + 142), radius=8, fill="#FFFFFF", outline=f"#{GRID}", width=2)
            draw.rectangle((x, y, x + 10, y + 142), fill=f"#{colours[index % len(colours)]}")
            item_font, item_lines = fitted_wrapped_font(draw, item["title"], 500, 2, 22, 18, True)
            draw.multiline_text((x + 35, y + 18), "\n".join(item_lines), fill=f"#{INK}", font=item_font, spacing=3)
            item_line_height = draw.textbbox((0, 0), "Ag", font=item_font)[3]
            purpose_y = y + 18 + len(item_lines) * item_line_height + max(0, len(item_lines) - 1) * 3 + 8
            purpose_font, purpose_lines = fitted_wrapped_font(
                draw,
                item["purpose"],
                500,
                3,
                16,
                11,
            )
            draw.multiline_text(
                (x + 35, purpose_y),
                "\n".join(purpose_lines),
                fill=f"#{MUTED}",
                font=purpose_font,
                spacing=3,
            )
    else:
        section_y = max(220, promise_bottom + 22)
        workbook_label = "workbook" if len(blueprint["catalogueItems"]) == 1 else "workbooks"
        draw.text((95, section_y), f"Inside the {workbook_label}", fill=f"#{GREEN}", font=font(24, True))
        if len(blueprint["catalogueItems"]) == 1:
            item = blueprint["catalogueItems"][0]
            card_top = min(300, max(274, section_y + 48))
            draw.rounded_rectangle((95, card_top, 1305, card_top + 180), radius=8, fill="#FFFFFF", outline=f"#{GRID}", width=2)
            draw.rectangle((95, card_top, 107, card_top + 180), fill=f"#{GREEN}")
            title_font, title_lines = fitted_wrapped_font(draw, item["title"], 1120, 2, 25, 19, True)
            draw.multiline_text((130, card_top + 22), "\n".join(title_lines), fill=f"#{INK}", font=title_font, spacing=3)
            title_height = draw.textbbox((0, 0), "Ag", font=title_font)[3]
            detail_y = card_top + 22 + len(title_lines) * title_height + max(0, len(title_lines) - 1) * 3 + 10
            draw_wrapped(draw, (130, detail_y), item["purpose"], f"#{MUTED}", font(17), 1120, 3, 4)

            status_count = len([column for column in item["columns"] if column.get("type") == "status"])
            metrics = [
                (str(len(item["columns"])), "named fields"),
                (str(status_count), "controlled dropdowns"),
                (str(len(item.get("calculations", []))), "live calculations"),
                (str(len(item["sampleRows"])), "example records"),
            ]
            metric_top = card_top + 210
            for index, (value, label) in enumerate(metrics):
                left = 95 + index * 307
                draw.rounded_rectangle((left, metric_top, left + 280, metric_top + 120), radius=8, fill="#FFFFFF", outline=f"#{GRID}", width=2)
                draw.text((left + 24, metric_top + 20), value, fill=f"#{GREEN}", font=font(30, True))
                draw.text((left + 24, metric_top + 67), label, fill=f"#{MUTED}", font=font(15, True))

            field_names = [safe_text(column["name"], 80) for column in item["columns"]]
            fields_top = metric_top + 154
            draw.text((95, fields_top), "Included fields", fill=f"#{INK}", font=font(18, True))
            draw_wrapped(
                draw,
                (95, fields_top + 34),
                " | ".join(field_names),
                f"#{MUTED}",
                font(14),
                1210,
                3,
                5,
            )
            headers = [safe_text(column["name"], 80) for column in item["columns"]]
            checks = sample_calculation_checks(item, calculation_definitions(item, headers))
            example_top = 744
            draw.text((95, example_top), "Example calculations", fill=f"#{INK}", font=font(18, True))
            for index, check in enumerate(checks[:2]):
                details = " | ".join(
                    f"{result['target']}: {result['display']}"
                    for result in check["results"]
                )
                draw.text(
                    (95, example_top + 34 + index * 28),
                    f"{check['record']} - {details}",
                    fill=f"#{MUTED}",
                    font=font(14),
                )
            image.save(output_path, format="PNG", optimize=True)
            with Image.open(output_path) as reopened:
                reopened.verify()
            return
        card_top = min(310, max(274, section_y + 48))
        card_width = 570
        card_height = 158
        row_gap = 14
        col_gap = 50
        colours = [GREEN, CYAN, AMBER, "8C86F7", "EC6F91", "61C0BF"]
        for index, item in enumerate(blueprint["catalogueItems"][:6]):
            column_index = index % 2
            row_index = index // 2
            left = 95 + column_index * (card_width + col_gap)
            top = card_top + row_index * (card_height + row_gap)
            right = left + card_width
            bottom = top + card_height
            draw.rounded_rectangle((left, top, right, bottom), radius=8, fill="#FFFFFF", outline=f"#{GRID}", width=2)
            draw.rectangle((left, top, left + 10, bottom), fill=f"#{colours[index % len(colours)]}")
            title_font, title_lines = fitted_wrapped_font(draw, item["title"], 505, 2, 19, 15, True)
            draw.multiline_text((left + 28, top + 14), "\n".join(title_lines), fill=f"#{INK}", font=title_font, spacing=2)
            title_height = draw.textbbox((0, 0), "Ag", font=title_font)[3]
            detail_y = top + 14 + len(title_lines) * title_height + max(0, len(title_lines) - 1) * 2 + 7
            field_names = [safe_text(column["name"], 80) for column in item["columns"]]
            fields_text = f"Fields: {', '.join(field_names[:4])}"
            if len(field_names) > 4:
                fields_text += f" +{len(field_names) - 4} more"
            detail_y = draw_wrapped(draw, (left + 28, detail_y), fields_text, f"#{MUTED}", font(13), 505, 2, 2) + 4
            status_column = next((
                column for column in item["columns"]
                if column.get("type") == "status" and re.search(r"(?:^|\s)status$", safe_text(column.get("name"), 80), flags=re.IGNORECASE)
            ), None)
            if status_column:
                options = [safe_text(value, 80) for value in status_column.get("options", [])]
                counted = positive_status(options)
                status_text = f"Dashboard: {counted} records" if counted else f"Controlled field: {safe_text(status_column.get('name'), 80)}"
                if "Ready" in options and "Approved" in options:
                    status_text = "Dashboard counts Approved records. Ready awaits final review."
                draw_wrapped(draw, (left + 28, detail_y), status_text, f"#{INK}", font(12, True), 505, 2, 2)
        footer_y = card_top + 3 * (card_height + row_gap) + 4
        draw.text(
            (95, min(812, footer_y)),
            "Coverage preview derived from every workbook's real fields, status controls, and dashboard logic.",
            fill=f"#{MUTED}",
            font=font(16),
        )
    image.save(output_path, format="PNG", optimize=True)
    with Image.open(output_path) as reopened:
        reopened.verify()


def cell_display(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")
    if isinstance(value, date):
        return value.strftime("%d %b %Y")
    return safe_text(value, 120)


def workbook_quality_preview(workbook_path, output_path, validation):
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    image = Image.new("RGB", (1600, 1000), f"#{INK}")
    draw = ImageDraw.Draw(image)
    draw.text((70, 42), "Actual workbook inspection", fill=f"#{WHITE}", font=font(32, True))
    draw.text(
        (70, 88),
        f"{workbook_path.name} | rendered from the saved XLSX",
        fill="#B9C5D3",
        font=font(17),
    )

    dashboard = workbook["Dashboard"] if "Dashboard" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    tracker = workbook["Tracker"] if "Tracker" in workbook.sheetnames else workbook[workbook.sheetnames[-1]]

    draw.rounded_rectangle((70, 140, 1530, 395), radius=8, fill="#FFFFFF", outline=f"#{GRID}", width=2)
    draw.text((96, 164), f"{dashboard.title} sheet", fill=f"#{INK}", font=font(21, True))
    dashboard_rows = []
    for row in dashboard.iter_rows(min_row=1, max_row=min(8, dashboard.max_row), values_only=True):
        values = [cell_display(value) for value in row[:6]]
        if any(values):
            dashboard_rows.append(values)
    dashboard_expected = validation.get("dashboardExpectedResults", {})
    y = 208
    for row in dashboard_rows[:6]:
        populated = [value for value in row if value]
        if not populated:
            continue
        label = populated[0]
        detail = " | ".join(populated[1:]) or ""
        if label == "Records entered":
            detail = f"{dashboard_expected.get('recordsEntered', 0)} expected from the sample rows | {detail}"
        elif label == dashboard_expected.get("metricLabel"):
            detail = f"{dashboard_expected.get('metricValue', 0)} expected from the sample rows | {detail}"
        draw.text((106, y), label[:42], fill=f"#{INK}", font=font(15, True))
        draw.text((535, y), detail[:105], fill=f"#{MUTED}", font=font(14))
        draw.line((96, y + 27, 1504, y + 27), fill=f"#{PANEL}", width=1)
        y += 32

    draw.rounded_rectangle((70, 425, 1530, 925), radius=8, fill="#FFFFFF", outline=f"#{GRID}", width=2)
    draw.text((96, 449), f"{tracker.title} sheet | actual headers and complete sample values", fill=f"#{INK}", font=font(21, True))
    max_columns = min(18, tracker.max_column)
    left = 96
    table_width = 1408
    rows_to_render = min(
        1 + max(1, len(validation.get("sampleCalculationChecks", []))),
        tracker.max_row,
    )
    sample_checks = {
        int(check["trackerRow"]): {
            result["target"]: result["display"]
            for result in check.get("results", [])
        }
        for check in validation.get("sampleCalculationChecks", [])
    }

    def draw_segment(start_column, end_column, top, label):
        column_count = end_column - start_column + 1
        column_width = table_width / column_count
        row_height = 54
        draw.text((left, top - 24), label, fill=f"#{MUTED}", font=font(13, True))
        for row_index in range(1, rows_to_render + 1):
            for offset, column_index in enumerate(range(start_column, end_column + 1)):
                x0 = int(left + offset * column_width)
                x1 = int(left + (offset + 1) * column_width)
                y0 = top + (row_index - 1) * row_height
                y1 = y0 + row_height
                fill = f"#{INK}" if row_index == 1 else ("#F7F9FC" if row_index % 2 == 0 else "#FFFFFF")
                text_colour = f"#{WHITE}" if row_index == 1 else f"#{INK}"
                draw.rectangle((x0, y0, x1, y1), fill=fill, outline=f"#{GRID}", width=1)
                header = cell_display(tracker.cell(row=1, column=column_index).value)
                value = cell_display(tracker.cell(row=row_index, column=column_index).value)
                if row_index > 1 and header in sample_checks.get(row_index, {}):
                    value = f"{sample_checks[row_index][header]} (checked)"
                shortened = textwrap.shorten(
                    value,
                    width=34 if row_index == 1 else 64,
                    placeholder="...",
                )
                wrapped = textwrap.wrap(shortened, width=20 if row_index == 1 else 24)[:3]
                draw.multiline_text(
                    (x0 + 7, y0 + 8),
                    "\n".join(wrapped),
                    fill=text_colour,
                    font=font(11 if row_index == 1 else 10, row_index == 1),
                    spacing=2,
                )

    split = min(8, max_columns)
    draw_segment(1, split, 500, "Operating record")
    if max_columns > split:
        draw_segment(split + 1, max_columns, 692, "Economics and notes")

    calculation_summary = []
    for check in validation.get("sampleCalculationChecks", []):
        values = " | ".join(
            f"{result['target']}: {result['display']}"
            for result in check.get("results", [])
        )
        calculation_summary.append(f"{check['record']} - {values}")
    draw.text(
        (96, 842),
        "Independent sample recalculation",
        fill=f"#{INK}",
        font=font(14, True),
    )
    for index, summary in enumerate(calculation_summary[:2]):
        draw.text(
            (96, 864 + index * 18),
            summary[:142],
            fill=f"#{MUTED}",
            font=font(11),
        )
    draw.text(
        (96, 902),
        f"Saved workbook facts: {tracker.max_column} fields | {tracker.max_row - 1} working rows | sheets: {', '.join(workbook.sheetnames)}",
        fill=f"#{MUTED}",
        font=font(13),
    )
    workbook.close()
    image.save(output_path, format="PNG", optimize=True)
    with Image.open(output_path) as reopened:
        reopened.verify()


def pdf_quality_preview(pdf_path, output_path, output_root=None):
    document = pdfium.PdfDocument(str(pdf_path))
    page_count = len(document)
    if page_count < 1:
        raise ValueError("The setup guide PDF has no pages")
    columns = 2 if page_count <= 2 else 3
    gap = 34
    side_margin = 70
    page_max_width = (
        1600 - (side_margin * 2) - (gap * (columns - 1))
    ) // columns - 16
    page_max_height = 790 if columns == 2 else 650
    page_images = []
    page_identities = []
    for page_index in range(page_count):
        page = document[page_index]
        bitmap = page.render(scale=1.45)
        rendered = bitmap.to_pil().convert("RGB")
        rendered.thumbnail((page_max_width, page_max_height), Image.Resampling.LANCZOS)
        page_image = rendered.copy()
        page_images.append(page_image)
        page_identities.append({
            "pageNumber": page_index + 1,
            "width": page_image.width,
            "height": page_image.height,
            "rasterSha256": hashlib.sha256(page_image.tobytes()).hexdigest(),
        })
        rendered.close()
        bitmap.close()
        page.close()
    document.close()

    rows = math.ceil(page_count / columns)
    row_height = page_max_height + 58
    image_height = 140 + (rows * row_height) + 40
    image = Image.new("RGB", (1600, image_height), f"#{INK}")
    draw = ImageDraw.Draw(image)
    draw.text((70, 42), "Actual setup guide inspection", fill=f"#{WHITE}", font=font(32, True))
    draw.text(
        (70, 88),
        f"{pdf_path.name} | {page_count} pages | rasterized from the saved PDF",
        fill="#B9C5D3",
        font=font(17),
    )
    for index, page_image in enumerate(page_images):
        row = index // columns
        column = index % columns
        row_images = page_images[row * columns:(row + 1) * columns]
        row_width = sum(item.width for item in row_images) + gap * max(0, len(row_images) - 1)
        row_x = (1600 - row_width) // 2
        x = row_x + sum(item.width for item in row_images[:column]) + gap * column
        row_top = 140 + row * row_height
        y = row_top + max(0, (page_max_height - page_image.height) // 2)
        draw.rounded_rectangle(
            (x - 8, y - 8, x + page_image.width + 8, y + page_image.height + 8),
            radius=5,
            fill="#DDE3EA",
        )
        image.paste(page_image, (x, y))
        draw.text(
            (x, row_top + page_max_height + 16),
            f"Page {index + 1}",
            fill="#B9C5D3",
            font=font(14, True),
        )
        page_image.close()
    image.save(output_path, format="PNG", optimize=True)
    with Image.open(output_path) as reopened:
        reopened.verify()
    evidence_root = Path(output_root or pdf_path.parent.parent).resolve()
    source_relative_path = pdf_path.resolve().relative_to(evidence_root).as_posix()
    inspection_relative_path = output_path.resolve().relative_to(evidence_root).as_posix()
    ordered_identity = json.dumps(
        page_identities,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
    )
    return {
        "sourceFile": pdf_path.name,
        "sourceRelativePath": source_relative_path,
        "sourceSha256": hashlib.sha256(pdf_path.read_bytes()).hexdigest(),
        "inspectionFile": output_path.name,
        "inspectionRelativePath": inspection_relative_path,
        "inspectionSha256": hashlib.sha256(output_path.read_bytes()).hexdigest(),
        "sourcePageCount": page_count,
        "renderedPageCount": len(page_images),
        "completeCoverage": len(page_images) == page_count,
        "columns": columns,
        "rows": rows,
        "pages": page_identities,
        "orderedPageIdentitySha256": hashlib.sha256(
            ordered_identity.encode("ascii")
        ).hexdigest(),
    }


def sha256(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def build(payload, output_root):
    spec = payload["spec"]
    manifest_filename = safe_output_leaf(
        spec.get("manifestFilename"),
        "manifestFilename",
        ".json",
    )
    bundle_filename = safe_output_leaf(
        spec.get("bundleFilename"),
        "bundleFilename",
        ".zip",
    )
    source_blueprint = payload["blueprint"]
    runtime_normalizations = payload.get("runtimeNormalizations", [])
    normalizations_by_item = {}
    for normalization in runtime_normalizations:
        item_id = str(normalization.get("itemId", ""))
        normalizations_by_item.setdefault(item_id, []).append({
            "code": safe_text(normalization.get("code"), 120),
            "reason": safe_text(normalization.get("reason"), 700),
        })
    safe_items = []
    item_adjustments = {}
    for source_item in source_blueprint["catalogueItems"]:
        safe_item, structure_adjustments = truth_aligned_item_structure(source_item)
        purpose, purpose_adjustments = claim_safe_item_purpose(safe_item)
        safe_item = {**safe_item, "purpose": purpose}
        safe_items.append(safe_item)
        item_adjustments[str(source_item["id"])] = [
            *normalizations_by_item.get(str(source_item["id"]), []),
            *structure_adjustments,
            *purpose_adjustments,
        ]
    source_setup_steps = [
        safe_text(step, 700)
        for step in source_blueprint.get("setupSteps", [])
        if safe_text(step, 700)
    ]
    setup_steps = canonical_package_setup_steps(safe_items)
    package_adjustments = []
    if source_setup_steps != setup_steps:
        package_adjustments.append({
            "code": "package_navigation_generated_from_runtime",
            "reason": "Pantheon generated package setup steps from the bundle and workbook structure it actually created.",
        })
    blueprint = {
        **source_blueprint,
        "customerPromise": claim_safe_copy(source_blueprint["customerPromise"]),
        "setupSteps": setup_steps,
        "catalogueItems": safe_items,
    }
    customer_root = output_root / "customer-files"
    preview_root = output_root / "storefront-previews"
    quality_review_root = output_root / "quality-review"
    customer_root.mkdir(parents=True, exist_ok=True)
    preview_root.mkdir(parents=True, exist_ok=True)
    quality_review_root.mkdir(parents=True, exist_ok=True)
    expected = {str(item["id"]) for item in spec["catalogueItems"]}
    actual = {str(item["id"]) for item in blueprint["catalogueItems"]}
    if expected != actual:
        raise ValueError("Blueprint catalogue IDs do not match the approved build specification")

    customer_files = {}
    customer_archive_paths = []
    validation = {}
    rendered_items = []
    for index, item in enumerate(blueprint["catalogueItems"], start=1):
        base = f"{index:02d}-{slug(item['title'])}"
        workbook_path = safe_contained_output_leaf(
            customer_root,
            f"{base}.xlsx",
            "workbookFilename",
            ".xlsx",
        )
        csv_path = safe_contained_output_leaf(
            customer_root,
            f"{base}-sample.csv",
            "sampleCsvFilename",
            ".csv",
        )
        instructions, instruction_adjustments = customer_instructions(
            item,
            workbook_path.name,
            csv_path.name,
        )
        item_adjustments[str(item["id"])].extend(instruction_adjustments)
        rendered_item = {
            **item,
            "columns": [
                {
                    **column,
                    "guidance": clarified_column_guidance(column) if column["type"] == "status" else (
                        "Choose Yes when this condition is true; otherwise choose No."
                        if column["type"] == "boolean"
                        else column["guidance"]
                    ),
                }
                for column in item["columns"]
            ],
            "instructions": instructions,
        }
        validation[item["id"]] = workbook_for_item(
            rendered_item,
            blueprint["packageTitle"],
            blueprint["customerPromise"],
            blueprint["setupSteps"],
            workbook_path,
        )
        write_csv(rendered_item, csv_path)
        customer_files[item["id"]] = [
            workbook_path.relative_to(output_root).as_posix(),
            csv_path.relative_to(output_root).as_posix(),
        ]
        customer_archive_paths.extend([workbook_path, csv_path])
        rendered_items.append(rendered_item)

    rendered_blueprint = {**blueprint, "catalogueItems": rendered_items}
    guide_path = safe_contained_output_leaf(
        customer_root,
        "00-customer-setup-guide.pdf",
        "setupGuideFilename",
        ".pdf",
    )
    guide_pdf(rendered_blueprint, customer_files, guide_path)
    preview_one = safe_contained_output_leaf(
        preview_root,
        "catalogue-overview.png",
        "cataloguePreviewFilename",
        ".png",
    )
    preview_two = safe_contained_output_leaf(
        preview_root,
        "workbook-preview.png",
        "workbookPreviewFilename",
        ".png",
    )
    preview_image(rendered_blueprint, preview_one, "dashboard")
    preview_image(rendered_blueprint, preview_two, "workbook")
    quality_workbook = safe_contained_output_leaf(
        quality_review_root,
        "actual-workbook.png",
        "workbookReviewFilename",
        ".png",
    )
    quality_guide = safe_contained_output_leaf(
        quality_review_root,
        "actual-setup-guide.png",
        "guideReviewFilename",
        ".png",
    )
    workbook_quality_preview(
        output_root / customer_files[rendered_items[0]["id"]][0],
        quality_workbook,
        validation[rendered_items[0]["id"]],
    )
    guide_inspection = pdf_quality_preview(guide_path, quality_guide, output_root)
    write_json(
        safe_contained_output_leaf(
            quality_review_root,
            "inspection-metadata.json",
            "inspectionMetadataFilename",
            ".json",
        ),
        {
            "schema": "pantheon.local-quality-inspection.v1",
            "images": {
                quality_guide.name: guide_inspection,
            },
        },
    )

    manifest = {
        "schema": "pantheon.product-manifest.v1",
        "version": 1,
        "planId": spec["planId"],
        "opportunityId": spec["opportunityId"],
        "packageTitle": blueprint["packageTitle"],
        "customerPromise": blueprint["customerPromise"],
        "runtimeNormalizations": runtime_normalizations,
        "catalogueItems": [
            {
                "id": item["id"],
                "title": item["title"],
                "purpose": item["purpose"],
                "files": customer_files[item["id"]],
                "validation": validation[item["id"]],
                "factoryAdjustments": item_adjustments[str(item["id"])],
            }
            for item in blueprint["catalogueItems"]
        ],
        "sharedFiles": [guide_path.relative_to(output_root).as_posix()],
        "setupGuide": {
            "path": guide_path.relative_to(output_root).as_posix(),
            "contentSource": "same_claim_safe_blueprint_used_to_render_pdf",
            "quickStart": [safe_text(step, 700) for step in rendered_blueprint["setupSteps"]],
            "products": [
                {
                    "title": safe_text(item["title"], 180),
                    "purpose": safe_text(item["purpose"], 1000),
                    "instructions": [safe_text(instruction, 700) for instruction in item["instructions"]],
                    "fields": [
                        [safe_text(column["name"], 100), safe_text(column["guidance"], 500)]
                        for column in item["columns"]
                    ],
                }
                for item in rendered_blueprint["catalogueItems"]
            ],
            "disclaimers": [safe_text(value, 800) for value in rendered_blueprint["disclaimers"]],
        },
        "storefrontPreviews": [
            preview_one.relative_to(output_root).as_posix(),
            preview_two.relative_to(output_root).as_posix(),
        ],
        "externalActionsTaken": [],
        "publishingStatus": "not_published",
        "factory": "pantheon-local-digital-product-factory-v1",
        "assetProvenance": {
            "sourceType": "pantheon_local_generation",
            "externalAssetsUsed": [],
            "customerDataUsed": False,
            "statement": (
                "Pantheon created the workbook, CSV, PDF, and PNG files locally from the approved "
                "product blueprint. No third-party image, logo, trademark, customer record, or "
                "downloaded design asset is embedded in the package."
            ),
        },
        "deliveryFormat": (
            f"{len(rendered_items)} editable Excel {'workbook' if len(rendered_items) == 1 else 'workbooks'}, "
            f"{len(rendered_items)} sample CSV {'file' if len(rendered_items) == 1 else 'files'}, "
            "one shared setup guide, and two storefront previews derived from the complete catalogue."
        ),
        "customerInstructionSource": "generated_from_actual_files_by_local_factory",
        "factoryAdjustments": package_adjustments,
    }

    archive_paths = sorted(
        [
            *customer_archive_paths,
            guide_path,
            preview_one,
            preview_two,
        ],
        key=lambda candidate: candidate.relative_to(output_root).as_posix(),
    )
    manifest["files"] = [
        {
            "path": file_path.relative_to(output_root).as_posix(),
            "bytes": file_path.stat().st_size,
            "sha256": sha256(file_path),
        }
        for file_path in archive_paths
    ]
    bundle_path = contained_output_leaf(
        output_root,
        bundle_filename,
        "bundleFilename",
    )
    manifest["bundle"] = {
        "filename": bundle_path.name,
        "canonicalManifestInsideBundle": True,
    }
    manifest_inside = contained_output_leaf(
        output_root,
        manifest_filename,
        "manifestFilename",
    )
    write_json(manifest_inside, manifest)
    archive_paths.append(manifest_inside)
    with zipfile.ZipFile(bundle_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for file_path in sorted(archive_paths, key=lambda candidate: candidate.relative_to(output_root).as_posix()):
            archive.writestr(
                canonical_zip_info(file_path.relative_to(output_root).as_posix()),
                file_path.read_bytes(),
            )

    with zipfile.ZipFile(bundle_path, "r") as archive:
        names = set(archive.namelist())
        for item in manifest["catalogueItems"]:
            for filename in item["files"]:
                if filename not in names:
                    raise ValueError(f"Bundle is missing {filename}")
        for filename in manifest["storefrontPreviews"]:
            if filename not in names:
                raise ValueError(f"Bundle is missing {filename}")
        embedded_manifest = archive.read(manifest_inside.name)
        if embedded_manifest != manifest_inside.read_bytes():
            raise ValueError("Bundle manifest is not identical to the standalone manifest")
        if archive.testzip() is not None:
            raise ValueError("Bundle integrity check failed")


def main():
    if len(sys.argv) != 3:
        raise SystemExit("usage: render-digital-product-kit.py INPUT_JSON OUTPUT_DIRECTORY")
    input_path = Path(sys.argv[1]).resolve()
    output_root = Path(sys.argv[2]).resolve()
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    if payload.get("schema") != "pantheon.digital-product-factory-input.v1":
        raise ValueError("Unsupported digital-product factory input schema")
    output_root.parent.mkdir(parents=True, exist_ok=True)
    if (
        output_root.is_symlink()
        or (output_root.exists() and not output_root.is_dir())
        or (output_root.exists() and any(output_root.iterdir()))
    ):
        raise ValueError(
            "Renderer output directory must be absent or empty before atomic commit"
        )
    stage_root = Path(tempfile.mkdtemp(
        prefix=".pantheon-render-",
        suffix=".tmp",
        dir=output_root.parent,
    )).resolve()
    try:
        build(payload, stage_root)
        if output_root.exists():
            output_root.rmdir()
        stage_root.replace(output_root)
    finally:
        if stage_root.exists():
            shutil.rmtree(stage_root)


if __name__ == "__main__":
    main()
